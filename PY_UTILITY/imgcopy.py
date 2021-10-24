import os
import sys
import re
import datetime
import threading
from distutils.dir_util import copy_tree
from collections import defaultdict
from openpyxl import load_workbook

threadList = list()
semaPhore = threading.Semaphore(5)

def fileSearching(folderList):
    global pidSet
    global fileData
    global threadList

    for folder in folderList:
        for day in list(fileData.keys()):
            curr = os.path.join(folder, day)
            pid_list = [str(pid[1]) for pid in fileData[day]]
        
            if not os.path.exists(curr):
                print(curr, '.....None Path')
                continue
            
            midd = os.path.join(resultPath, day)
            
            if not os.path.exists(midd):
                try:
                    os.mkdir(midd)

                except Exception as e:
                    print(e)
            
            for root, _, files in os.walk(curr):
                for file in files:
                    fileProcessing(pid_list, day, midd, root, file)

            currThreads = list(threadList)
            for thread in currThreads:
                thread.join()

            threadList = list(set(threadList) ^ set(currThreads))


def fileProcessing(pidList, day, middleName, dirName, fileName):
    reg_obj = patt.search(fileName)               
    if reg_obj and len(reg_obj.groups()[-1]) <= 8:
        pid = str(int(reg_obj.groups()[-1])).strip()
                            
        if pid in pidList:
            idx = fileData[day][pidList.index(pid)][0]
                                
            if not os.path.exists(os.path.join(middleName, f"{idx}_{pid}")):
                os.mkdir(os.path.join(middleName, f"{idx}_{pid}"))
                                    
            else:
                if pid not in pidSet:
                    pidSet.add(pid)
                    print(f"pid : {pid}")
                    return
                                    
                else:
                    return
                                    
            try:
                if not os.listdir(os.path.join(middleName, f"{idx}_{pid}")):
                    semaPhore.acquire()

                    parallel = threading.Thread(
                        target=copy_tree,
                        args=(dirName, os.path.join(middleName, f"{idx}_{pid}")))
                    parallel.daemon = True
                    parallel.start()

                    threadList.append(parallel)
                    semaPhore.release()
                    
            except Exception as e:
                print(e)

def noneImageProcessing(noneImgDict):
    logFilePath = r'C:\Cognex\LMS\Log'
    CB = ['BCR1', 'BCR4']

    for day in noneImgDict.keys():
        for area in noneImgDict[day]:
            path = os.path.join(logFilePath, area, day + '.txt')
            if not os.path.exists(path):
                path = os.path.join(logFilePath, area, day + '.csv')
            
            logContents = open(
                file=path,
                mode='r',
                encoding='euc-kr',
                errors='replace').readlines()

            currPIDs = list()

            for line in logContents:
                if area in CB:
                    pattObject = re.compile(r"D,2,(\d{8})")

                else:
                    pattObject = re.compile(r"D,(\d{6})")

                cmpObject = pattObject.search(line)
                if not cmpObject:
                    continue

                currPIDs.append(str(int(cmpObject.groups()[0])).strip())

            for pid in noneImgDict[day][area]:
                idx = None
                try:
                    idx = currPIDs.index(pid)
                
                except ValueError:
                    print(f"pid : {pid}가 로그에 존재하지 않습니다.")
                    continue
                
                print("pid :", pid, "// index :", idx)

                areaPath = f"D:\\{area}-2\\{day}"
                idxFileList = os.listdir(areaPath)
                currentPath = None

                trueFlag = True
                offset = 0
                count = 0
                currPID = 0

                while trueFlag:

                    if idx > (len(idxFileList) / 2):
                        count = -(count)
                    
                    if not offset:
                        try:
                            currentPath = os.path.join(areaPath, idxFileList[idx + count])

                        except IndexError:
                            print(f"currentPath[{idx}]는 파일 항목 수량({len(idxFileList)})을 초과합니다.")
                            sys.exit()

                    count += 1
                
                    for fileName in os.listdir(currentPath):
                        pattCmp = patt.search(fileName)
                        if pattCmp and len(pattCmp.groups()[-1]) <= 8:
                            currPID = str(int(pattCmp.groups()[-1])).strip()

                            try:
                                currPIDs.index(currPID)
                                
                            except ValueError:
                                print(f"currPIDs.index({currPID}) Value Error")
                                continue

                
                            else:
                                if idx > currPIDs.index(currPID):
                                    offset = idx + (idx - currPIDs.index(currPID))

                                else:
                                    offset = idx - currPIDs.index(currPID)
   
                                
                        if os.path.exists(os.path.join(areaPath, idxFileList[offset])):
                            trueFlag = False
                
                #noneImagePIDPath = os.path.join(noneImageFolder, pid)

                #if offset:
                 #   if not os.path.exists(noneImagePIDPath):
                 #       os.mkdir(os.path.join(noneImageFolder, pid))
                 #       copy_tree(currentPath, os.path.join(noneImageFolder, pid))
                 #       print("pid : " + str(pid) + f", Log 위치({idx})에서 offset({offset})만큼 이동한 위치의 파일 복사")

                #else:
                #    print(pid, ": 폴더을 찾지 못함")

semaPhore = threading.Semaphore(3)
excelPath = input("Excel 주소\n>>>")
if not os.path.exists(excelPath):
    sys.exit()

drivePath = None
if not os.path.exists("D:\\"):
    drivePath = input("D 드라이브가 없습니다. 저장할 새 경로를 입력해주세요.\n>>>")

    if not os.path.exists(drivePath):
        if os.path.exists(os.path.dirname(drivePath)):
            try:
                os.mkdir(drivePath)
            
            except Exception as e:
                print(e)
                os.system('pause')
                sys.exit()
        else:   
            print("경로가 잘못되었습니다.\n다시 실행해주세요.")
            os.system('pause')
            sys.exit()

    elif os.path.isfile(drivePath):
        print("경로가 파일입니다.\n폴더로 입력해주세요.")
        os.system('pause')
        sys.exit()

else:
    drivePath = "D:\\"
    
patt = re.compile(r"^#(\d+)#(\d+)")
fileData = defaultdict(list)
excelObject = load_workbook(excelPath)
sheet = excelObject.active
# pid = row 2 / column 1
# scan_dt = row 2 / column 3

def timeCasting(year, month, day):
    m, d= str(month), str(day)

    if len(m) < 2:
        m = '0' + m

    if len(d) < 2:
        d = '0' + d

    return ''.join([str(year), m, d])

def machineSort(pid):
    pid = int(pid)
    area = None

    if pid < 20000:
        area = "W1"

    elif pid < 30000:
        area = "W2"

    else:
        area = "BCR4"

    return area


for idx, row in enumerate(range(2, 1001), 2):
    
    pid = str(sheet.cell(row, 1).value)
    scan_dt = str(sheet.cell(row, 3).value)
    
    if pid == '' or pid is None or \
       scan_dt == '' or scan_dt is None:
        continue
    
    fileData[scan_dt[:8]].append((idx, pid))


timeObject = datetime.datetime.now()
subFilePath = timeCasting(timeObject.year,
                          timeObject.month,
                          timeObject.day)

file_cnt = 1
resultPath = os.path.join(drivePath, subFilePath + f"_{file_cnt}")

if not os.path.exists(resultPath):
    os.mkdir(resultPath)
    
else:
    while os.path.exists(resultPath):
        file_cnt += 1
        resultPath = os.path.join(drivePath, subFilePath + f"_{file_cnt}")

    os.mkdir(resultPath)


dsk_address = os.path.join(os.path.expanduser('~'), 'Desktop')
imgFolderList = [os.path.join(dsk_address, pp) for pp in ['BCR1-2', 'BCR4-2', 'W1-2', 'W2-2']]
print(imgFolderList)

pidSet = set()
fileSearching(imgFolderList)

print("PID가 존재하는 파일 복사 완료..")
#print("PID가 존재하지 않는 파일 검색 중..")

#noneFindPID = defaultdict(list)

#for day in os.listdir(resultPath):
#    noneFindPID[day] = defaultdict(list)
#    findData = {file.split('_')[-1] for file in os.listdir(os.path.join(resultPath, day))}#
#
#    for pid in ({file[-1] for file in fileData[day]} ^ findData):
#        noneFindPID[day][machineSort(pid)].append(pid)

#noneImageFolder = os.path.join(resultPath, "noneImages")
#if not os.path.exists(noneImageFolder):
#    os.mkdir(noneImageFolder)

#noneImageProcessing(noneFindPID)

#print("PID가 존재하지 않는 파일 복사 완료.")

os.system('pause')

                    

                    
                
                
                
                    
                    
                    
                
    
    
    
